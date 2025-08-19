from flask import Flask, render_template, request, jsonify, send_file, after_this_request
import io
import pandas as pd
import numpy as np
from typing import List
import tempfile
import os
from utils import spds_value_labels_map

app = Flask(__name__, static_folder="static", template_folder="templates")
app.config["MAX_CONTENT_LENGTH"] = 1024 * 1024 * 200  # até ~200MB

# Estado atual em memória (nada vai para disco)
_current_df = None
_current_meta = None
_current_filename = None
_last_tables = None  # guarda as últimas tabelas (para export)

# ----------------- helpers -----------------
ALLOWED_EXTENSIONS = {"sav"}

def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def build_user_missing_mask(meta, df: pd.DataFrame, vars_list: List[str]) -> pd.Series:
    """True = mantém linha; False = remove (user-missing do SPSS)."""
    if meta is None:
        return pd.Series(True, index=df.index)
    mv_map = getattr(meta, "missing_user_values", {}) or {}
    mr_map = getattr(meta, "missing_ranges", {}) or {}
    keep = pd.Series(True, index=df.index)
    for v in vars_list:
        if v not in df.columns:
            continue
        col = df[v]
        bad = pd.Series(False, index=df.index)
        vals = mv_map.get(v)
        if vals:
            bad |= col.isin(vals)
        ranges = mr_map.get(v)
        if ranges:
            tmpbad = pd.Series(False, index=df.index)
            for low, high in ranges:
                try:
                    cnum = pd.to_numeric(col, errors="coerce")
                    tmpbad |= (cnum >= float(low)) & (cnum <= float(high))
                except Exception:
                    pass
            bad |= tmpbad
        keep &= ~bad
    return keep

def var_label(meta, var_name: str) -> str:
    """Label amigável da variável; fallback para o próprio nome."""
    if meta is None:
        return var_name
    mapping = getattr(meta, "column_names_to_labels", {}) or {}
    return mapping.get(var_name, var_name)

# ----------------- rotas -----------------
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/upload", methods=["POST"])
def upload():
    global _current_df, _current_meta, _current_filename
    try:
        if "file" not in request.files:
            return jsonify({"error": "Nenhum arquivo enviado."}), 400
        f = request.files["file"]
        if f.filename == "":
            return jsonify({"error": "Nome de arquivo vazio."}), 400
        if not allowed_file(f.filename):
            return jsonify({"error": "Envie um arquivo .sav"}), 400

        import pyreadstat, tempfile, os
        # escreve os bytes do upload em um arquivo temporário
        raw = f.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".sav") as tmp:
            tmp.write(raw)
            tmp_path = tmp.name

        try:
            _current_df, _current_meta = pyreadstat.read_sav(tmp_path, apply_value_formats=False)
            _current_df.columns = [str(c) for c in _current_df.columns]
            _current_filename = f.filename
        finally:
            os.remove(tmp_path)  # apaga o temp imediatamente após a leitura

        vars_info = [{"name": c, "dtype": str(_current_df[c].dtype)} for c in _current_df.columns]
        suggested_weight = "peso" if "peso" in _current_df.columns else None

        return jsonify({
            "message": f"Arquivo carregado: {_current_filename}",
            "variables": vars_info,
            "suggested_weight": suggested_weight,
            "filename": _current_filename
        })
    except Exception as e:
        return jsonify({"error": f"Falha no upload: {e}"}), 500


@app.route("/close", methods=["POST"])
def close_file():
    """Limpa tudo da memória (não há nada no disco para apagar)."""
    global _current_df, _current_meta, _current_filename, _last_tables
    _current_df = None
    _current_meta = None
    _current_filename = None
    _last_tables = None
    return jsonify({"message": "Arquivo fechado e memória limpa."})

# ------------- crosstab (uma tabela por cruzamento) -------------
@app.route("/crosstab", methods=["POST"])
def crosstab():
    global _current_df, _current_meta, _last_tables
    if _current_df is None:
        return jsonify({"error": "Nenhum arquivo carregado."}), 400

    try:
        data = request.get_json(force=True)
        rows_in = data.get("rows") or []
        cols_in = data.get("cols") or []
        weight_var = data.get("weight", "peso")
        options = list(dict.fromkeys(data.get("options", [])))
        niw_mode = data.get("niw_mode", "none")

        if not rows_in or not cols_in:
            return jsonify({"error": "Selecione ao menos uma variável para linha e coluna."}), 400

        need_cols = list(dict.fromkeys(rows_in + cols_in + [weight_var]))
        base = _current_df[need_cols].copy()
        base[weight_var] = pd.to_numeric(base[weight_var], errors="coerce")
        base = base.dropna(subset=[weight_var])

        if niw_mode == "round_case":
            base[weight_var] = np.round(base[weight_var])
        elif niw_mode == "trunc_case":
            base[weight_var] = np.floor(base[weight_var])

        labels_map = spds_value_labels_map(_current_meta)

        def to_ordered_cat(series, var):
            labels_dict = labels_map.get(var)
            if labels_dict:
                codes = list(labels_dict.keys())
                try:
                    codes_sorted = sorted(codes, key=lambda x: float(x))
                except Exception:
                    codes_sorted = codes
            else:
                uniq = series.dropna().unique().tolist()
                try:
                    codes_sorted = sorted(uniq, key=lambda x: float(x))
                except Exception:
                    codes_sorted = sorted(map(str, uniq))
            if all(isinstance(k, (int, float)) for k in codes_sorted):
                ser = pd.to_numeric(series, errors="coerce")
            else:
                ser = series.astype(series.dtype)
            return pd.Categorical(ser, categories=codes_sorted, ordered=True), labels_dict

        def fmt_value(metric, v, mode):
            if pd.isna(v):
                return ""
            if metric in ("% Row", "% Column", "% Total"):
                return f"{v:.1f}%"
            if metric in ("Observed", "Expected"):
                return f"{v:.0f}" if mode in ("round_cell", "trunc_cell") else f"{v:.2f}"
            return f"{v:.1f}"

        def compute_one(row_var, col_var):
            df = base[[row_var, col_var, weight_var]].copy()
            df = df[build_user_missing_mask(_current_meta, df, [row_var, col_var])].copy()

            df[row_var], row_lbls = to_ordered_cat(df[row_var], row_var)
            df[col_var], col_lbls = to_ordered_cat(df[col_var], col_var)

            obs = pd.pivot_table(
                df,
                values=weight_var,
                index=[row_var],
                columns=[col_var],
                aggfunc="sum",
                fill_value=0.0,
                margins=True,
                margins_name="Total",
                observed=True,
            )
            if niw_mode == "round_cell":
                obs = np.round(obs)
            elif niw_mode == "trunc_cell":
                obs = np.floor(obs)

            core = obs.drop(index="Total", errors="ignore").drop(columns="Total", errors="ignore")
            row_tot = core.sum(axis=1)
            col_tot = core.sum(axis=0)
            grand_total = float(core.values.sum()) if core.size else 0.0

            mats = {}
            if "row_pct" in options:
                pct = core.div(row_tot.replace(0, np.nan), axis=0) * 100.0
                m = obs.copy()
                m.loc[core.index, core.columns] = pct
                mats["% Row"] = m
            if "col_pct" in options:
                pct = core.div(col_tot.replace(0, np.nan), axis=1) * 100.0
                m = obs.copy()
                m.loc[core.index, core.columns] = pct
                mats["% Column"] = m
            if "total_pct" in options:
                mats["% Total"] = (obs / grand_total) * 100.0 if grand_total > 0 else obs * 0

            if any(k in options for k in ["expected", "resid", "std_resid", "adj_std_resid"]):
                expected = (
                    np.outer(row_tot.values, col_tot.values) / grand_total if grand_total > 0 else np.zeros_like(core.values)
                )
                expected = pd.DataFrame(expected, index=core.index, columns=core.columns)
                resid = core - expected
                if "expected" in options:
                    m = obs.copy()
                    m.loc[core.index, core.columns] = expected
                    mats["Expected"] = m
                if "resid" in options:
                    m = obs.copy()
                    m.loc[core.index, core.columns] = resid
                    mats["Residual (O−E)"] = m
                with np.errstate(divide="ignore", invalid="ignore"):
                    std = resid / np.sqrt(expected.replace(0, np.nan))
                if "std_resid" in options:
                    m = obs.copy()
                    m.loc[core.index, core.columns] = std
                    mats["Standardized Residual"] = m
                if "adj_std_resid" in options:
                    if grand_total > 0:
                        p_row = (row_tot / grand_total).values[:, None]
                        p_col = (col_tot / grand_total).values[None, :]
                        denom = np.sqrt(expected.values * (1 - p_row) * (1 - p_col))
                        adj = resid.values / np.where(denom == 0, np.nan, denom)
                        adj = pd.DataFrame(adj, index=core.index, columns=core.columns)
                    else:
                        adj = core * 0
                    m = obs.copy()
                    m.loc[core.index, core.columns] = adj
                    mats["Adjusted Standardized Residual"] = m

            display = []
            if "observed" in options:
                display.append(("Observed", obs))
            if "expected" in options and "Expected" in mats:
                display.append(("Expected", mats["Expected"]))
            if "row_pct" in options and "% Row" in mats:
                display.append(("% Row", mats["% Row"]))
            if "col_pct" in options and "% Column" in mats:
                display.append(("% Column", mats["% Column"]))
            if "total_pct" in options and "% Total" in mats:
                display.append(("% Total", mats["% Total"]))
            if "resid" in options and "Residual (O−E)" in mats:
                display.append(("Residual (O−E)", mats["Residual (O−E)"]))
            if "std_resid" in options and "Standardized Residual" in mats:
                display.append(("Standardized Residual", mats["Standardized Residual"]))
            if "adj_std_resid" in options and "Adjusted Standardized Residual" in mats:
                display.append(("Adjusted Standardized Residual", mats["Adjusted Standardized Residual"]))

            idx_labels = [row_lbls.get(v, "Total" if v == "Total" else str(v)) for v in obs.index]
            col_labels = [col_lbls.get(v, "Total" if v == "Total" else str(v)) for v in obs.columns]

            data_cells = []
            flag_mask = []
            for r in range(obs.shape[0]):
                row_cells = []
                row_flags = []
                is_total_row = obs.index[r] == "Total"
                for c in range(obs.shape[1]):
                    is_total_col = obs.columns[c] == "Total"
                    lines = []
                    for name, mat in display:
                        v = mat.iat[r, c] if (r < mat.shape[0] and c < mat.shape[1]) else np.nan
                        lines.append(f"{name}: {fmt_value(name, v, niw_mode)}")
                    flagged = False
                    if not is_total_row and not is_total_col and "adj_std_resid" in options and "Adjusted Standardized Residual" in mats:
                        try:
                            v = mats["Adjusted Standardized Residual"].iat[r, c]
                            flagged = (abs(float(v)) > 1.9) if pd.notna(v) else False
                        except Exception:
                            flagged = False
                    row_cells.append(lines)
                    row_flags.append(flagged)
                data_cells.append(row_cells)
                flag_mask.append(row_flags)

            title_label = f"{var_label(_current_meta, row_var)} × {var_label(_current_meta, col_var)}"

            return {
                "title": title_label,
                "index": idx_labels,
                "columns": col_labels,
                "data": data_cells,
                "flagMask": flag_mask,
            }

        tables = {}
        for r in rows_in:
            for c in cols_in:
                key = f"[{var_label(_current_meta, r)} × {var_label(_current_meta, c)}]"
                tables[key] = compute_one(r, c)

        _last_tables = tables
        return jsonify({"tables": tables})
    except Exception as e:
        return jsonify({"error": f"Falha ao gerar a tabela: {e}"}), 500

# ------------- export excel (TEMPORÁRIO: apaga após envio) -------------
@app.route("/export_excel", methods=["GET"])
def export_excel():
    global _last_tables
    if not _last_tables:
        return jsonify({"error": "Nenhuma tabela gerada ainda."}), 400

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Tabelas"

    header_font = Font(bold=True)
    title_font = Font(bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    hl_fill = PatternFill(start_color="FFFDE68A", end_color="FFFDE68A", fill_type="solid")

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")

    def apply_grid(r1, c1, r2, c2, outline_medium=True):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=c)
                left_side = thin; right_side = thin; top_side = thin; bottom_side = thin
                if outline_medium:
                    if r == r1: top_side = medium
                    if r == r2: bottom_side = medium
                    if c == c1: left_side = medium
                    if c == c2: right_side = medium
                cell.border = Border(left=left_side, right=right_side, top=top_side, bottom=bottom_side)

    row_cursor = 1
    for _, t in _last_tables.items():
        ncols = len(t["columns"])
        first_col = 1
        last_col = 1 + ncols

        ws.cell(row=row_cursor, column=first_col, value=t["title"]).font = title_font
        ws.merge_cells(start_row=row_cursor, start_column=first_col, end_row=row_cursor, end_column=last_col)
        ws.cell(row=row_cursor, column=first_col).alignment = center
        apply_grid(row_cursor, first_col, row_cursor, last_col)
        row_cursor += 1

        header_row = row_cursor
        ws.cell(row=header_row, column=first_col, value="").font = header_font
        ws.cell(row=header_row, column=first_col).alignment = center
        for j, colname in enumerate(t["columns"], start=2):
            cell = ws.cell(row=header_row, column=j, value=colname)
            cell.font = header_font
            cell.alignment = center
        apply_grid(header_row, first_col, header_row, last_col)
        row_cursor += 1

        nrows = len(t["index"])
        for i in range(nrows):
            heights = []
            for j in range(ncols):
                cell_content = t["data"][i][j]
                h = len(cell_content) if isinstance(cell_content, list) else 1
                heights.append(h)
            block_height = max(heights) if heights else 1

            block_start = row_cursor
            block_end = row_cursor + block_height - 1

            ws.cell(row=block_start, column=first_col, value=t["index"][i]).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            if block_height > 1:
                ws.merge_cells(start_row=block_start, start_column=first_col, end_row=block_end, end_column=first_col)

            for j in range(ncols):
                lines = t["data"][i][j] if isinstance(t["data"][i][j], list) else [t["data"][i][j]]
                for k in range(block_height):
                    r = row_cursor + k
                    text = lines[k] if k < len(lines) and lines[k] is not None else ""
                    cell = ws.cell(row=r, column=j + 2, value=text)
                    cell.alignment = left_wrap

                    flagged = False
                    try:
                        flagged = bool(t.get("flagMask") and t["flagMask"][i][j])
                    except Exception:
                        flagged = False
                    if flagged and isinstance(text, str) and text.startswith("Adjusted Standardized Residual:"):
                        cell.fill = hl_fill

            apply_grid(block_start, first_col, block_end, last_col)
            row_cursor += block_height

        row_cursor += 1

        for c in range(first_col, last_col + 1):
            maxlen = 0
            for r in range(header_row, row_cursor):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                maxlen = max(maxlen, len(str(v)))
            ws.column_dimensions[get_column_letter(c)].width = min(45, max(12, int(maxlen * 0.95)))

    # ----- SALVA EM ARQUIVO TEMPORÁRIO E REMOVE APÓS RESPOSTA -----
    fd, tmp_path = tempfile.mkstemp(prefix="tabelas_", suffix=".xlsx")
    os.close(fd)
    wb.save(tmp_path)

    @after_this_request
    def _cleanup_export(response):
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        return response

    return send_file(tmp_path, as_attachment=True, download_name="tabelas.xlsx")

if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))  # Render injeta PORT
    app.run(host="0.0.0.0", port=port, debug=False)
